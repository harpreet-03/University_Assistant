"""
parsers.py
Reads every file type and returns plain text chunks for the vector DB.

DOCX parsing strategy (key insight from real LPU documents):
  - Many university DOCX files use NO heading styles — they use bold Normal text as headers
  - We detect bold paragraphs as section breaks
  - We chunk by SECTION (bold header + its content), not by word count alone
  - Each section becomes its own searchable chunk → accurate retrieval
  - Tables are kept with their surrounding section context

CHUNKING philosophy:
  - A chunk should be ONE complete idea (a policy, a step, a rule, a section)
  - Not a fixed word count slice that cuts through meaning
  - Overlap is added for long sections only
"""

import re
from pathlib import Path


# ═══════════════════════════════════════════════════════════════
#  XLSX  (Flat timetable format — all_timetables.xlsx)
#
#  Expected columns (row 1 = headers):
#    faculty_id | faculty_name | day | time_slot |
#    course_code | course_name | section | room | class_type
#
#  One row = one class slot. No merged cells. No cryptic encoding.
#  Each faculty gets one chunk per teaching day + one summary chunk.
# ═══════════════════════════════════════════════════════════════

_FLAT_HEADERS = [
    "faculty_id", "faculty_name", "day", "time_slot",
    "course_code", "course_name", "section", "room", "class_type",
]

_DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

def parse_xlsx_to_rows(filepath):
    """
    Parse flat-format xlsx -> list of dicts for direct SQL insertion.
    Each dict has keys: faculty_id, faculty_name, day, time_slot,
                        course_code, course_name, section, room, class_type
    Used by sqldb.insert_timetable() -- NOT the vector DB.
    """
    import openpyxl
    ws   = openpyxl.load_workbook(filepath).active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    first_row = [str(v).strip().lower() if v else "" for v in rows[0]]
    if "faculty_id" not in first_row or "course_code" not in first_row:
        return []

    col = {name: first_row.index(name)
           for name in _FLAT_HEADERS if name in first_row}

    def get(row, name):
        idx = col.get(name)
        return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] else ""

    result = []
    for row in rows[1:]:
        fid = get(row, "faculty_id")
        day = get(row, "day")
        if not fid or not day:
            continue
        result.append({
            "faculty_id":   fid,
            "faculty_name": get(row, "faculty_name"),
            "day":          day,
            "time_slot":    get(row, "time_slot"),
            "course_code":  get(row, "course_code"),
            "course_name":  get(row, "course_name"),
            "section":      get(row, "section"),
            "room":         get(row, "room"),
            "class_type":   get(row, "class_type"),
        })
    return result




def parse_xlsx(filepath):
    import openpyxl
    ws   = openpyxl.load_workbook(filepath).active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    # ── Detect format ─────────────────────────────────────────
    # Flat format: first row contains our known column headers
    first_row = [str(v).strip().lower() if v else "" for v in rows[0]]
    is_flat   = "faculty_id" in first_row and "course_code" in first_row

    if not is_flat:
        # Unrecognised format — return empty and warn
        return [
            "Timetable file could not be parsed. "
            "Please use the flat format with columns: "
            "faculty_id, faculty_name, day, time_slot, "
            "course_code, course_name, section, room, class_type."
        ]

    # ── Map column names to indices ───────────────────────────
    col = {name: first_row.index(name) for name in _FLAT_HEADERS if name in first_row}

    def get(row, name):
        idx = col.get(name)
        return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] else ""

    # ── Group rows by faculty then by day ─────────────────────
    # structure: {faculty_id: {day: [slot_dict, ...]}}
    faculty_info = {}   # faculty_id → faculty_name
    schedule     = {}   # faculty_id → {day → [slots]}
    specials     = {}   # faculty_id → [special slots]

    for row in rows[1:]:
        fid   = get(row, "faculty_id")
        fname = get(row, "faculty_name")
        day   = get(row, "day")
        if not fid or not day:
            continue

        faculty_info[fid] = fname

        slot = {
            "time":        get(row, "time_slot"),
            "course_code": get(row, "course_code"),
            "course_name": get(row, "course_name"),
            "section":     get(row, "section"),
            "room":        get(row, "room"),
            "class_type":  get(row, "class_type"),
        }

        if day.startswith("Special"):
            specials.setdefault(fid, []).append({**slot, "day": day})
        else:
            schedule.setdefault(fid, {}).setdefault(day, []).append(slot)

    # ── Build chunks ──────────────────────────────────────────
    chunks = []

    for fid, fname in sorted(faculty_info.items()):
        fac_days    = schedule.get(fid, {})
        fac_special = specials.get(fid, [])

        # Collect all courses this faculty teaches
        all_courses = sorted({
            f"{s['course_code']} ({s['course_name']})"
            for day_slots in fac_days.values()
            for s in day_slots
            if s["course_code"]
        })
        active_days = [d for d in _DAY_ORDER if d in fac_days]

        # Summary chunk — answers "who teaches what / when"
        chunks.append(
            f"Faculty: {fname}, ID: {fid}. "
            f"Teaching days: {', '.join(active_days)}. "
            f"Courses: {', '.join(all_courses) if all_courses else 'N/A'}."
        )

        # One chunk per teaching day
        for day in _DAY_ORDER:
            slots = fac_days.get(day)
            if not slots:
                continue
            lines = [f"Timetable of {fname} (ID: {fid}) on {day}:"]
            for s in slots:
                course_label = (
                    f"{s['course_code']} - {s['course_name']}"
                    if s["course_name"] and s["course_name"] != s["course_code"]
                    else s["course_code"]
                )
                lines.append(
                    f"  {s['time']}  |  Course: {course_label}  |  "
                    f"Section: {s['section'] or '?'}  |  "
                    f"Room: {s['room'] or '?'}  |  {s['class_type'] or 'Lecture'}"
                )
            chunks.append("\n".join(lines))

        # Special / activity classes chunk
        if fac_special:
            lines = [f"Special / activity classes for {fname} (ID: {fid}):"]
            for s in sorted(fac_special, key=lambda x: x["day"]):
                lines.append(
                    f"  {s['day']}  {s['time']}  |  "
                    f"Course: {s['course_code']}  |  "
                    f"Section: {s['section']}  |  Room: {s['room']}"
                )
            chunks.append("\n".join(lines))

    return chunks


# ═══════════════════════════════════════════════════════════════
#  DOCX  — Robust, section-aware parser
# ═══════════════════════════════════════════════════════════════

def _is_section_header(p):
    """
    Detect if a paragraph is a section header.
    University docs often use bold Normal text instead of Heading styles.
    """
    style = p.style.name if p.style else "Normal"
    text  = p.text.strip()
    if not text:
        return False

    # Explicit heading styles
    if "Heading" in style or style in ("Title", "Subtitle"):
        return True

    # Bold paragraph = section header (very common in LPU docs)
    # Check ALL runs — if at least one significant run is bold, treat as header
    runs_with_text = [r for r in p.runs if r.text.strip()]
    if runs_with_text and all(r.bold for r in runs_with_text):
        return True

    # ALL CAPS text of reasonable length = section header
    if text == text.upper() and len(text.split()) >= 3:
        return True

    return False


def _table_to_text(tbl, doc):
    """Convert table to text, handling merged cells via XML element deduplication."""
    from docx.table import Table as DocxTable

    rows_text = []
    for row in tbl.rows:
        seen = set()
        cells = []
        for cell in row.cells:
            cid = id(cell._tc)
            if cid in seen:
                continue
            seen.add(cid)
            # Recursively extract cell content (handles nested tables)
            cell_parts = []
            for child in cell._tc:
                ctag = child.tag.split("}")[-1]
                if ctag == "p":
                    from docx.text.paragraph import Paragraph
                    cp = Paragraph(child, doc)
                    ct = cp.text.strip()
                    if ct:
                        cell_parts.append(ct)
                elif ctag == "tbl":
                    nested = DocxTable(child, doc)
                    nt = _table_to_text(nested, doc)
                    if nt:
                        cell_parts.append(nt)
            joined = " ".join(cell_parts)
            if joined:
                cells.append(joined)
        if cells:
            rows_text.append(" | ".join(cells))
    return "\n".join(rows_text)


def parse_docx(filepath):
    """
    Section-aware DOCX parser.

    Strategy:
    1. Walk body elements in order (paragraphs + tables interleaved)
    2. Treat bold/heading paragraphs as section boundaries
    3. Accumulate each section as ONE chunk (header + its content)
    4. If a section is too long (>300 words), split it with overlap
    5. Tables inherit the section they belong to

    This ensures each chunk is a semantically complete unit
    (one policy, one SOP step, one scholarship scheme) — not a
    random word-count slice that bleeds across topics.
    """
    from docx import Document
    from docx.text.paragraph import Paragraph
    from docx.table import Table as DocxTable

    doc = Document(filepath)

    # ── Step 1: Extract elements in document order ─────────────
    # Each element is ("text", text_string) or ("header", header_string)
    elements = []

    for child in doc.element.body:
        tag = child.tag.split("}")[-1]

        if tag == "p":
            p = Paragraph(child, doc)
            text = "".join(run.text for run in p.runs).strip()
            if not text:
                text = p.text.strip()
            if not text:
                continue

            style = p.style.name if p.style else "Normal"

            # Bullet / list items
            if re.match(r"List (Bullet|Number|Paragraph)", style):
                prefix = "• " if "Bullet" in style or "Paragraph" in style else "- "
                elements.append(("text", prefix + text))

            elif _is_section_header(p):
                elements.append(("header", text))

            else:
                elements.append(("text", text))

        elif tag == "tbl":
            tbl = DocxTable(child, doc)
            table_text = _table_to_text(tbl, doc)
            if table_text:
                elements.append(("text", "[TABLE]\n" + table_text))

    # ── Step 2: Group elements into sections ───────────────────
    # A section = one header + everything until the next header
    sections = []
    current_header  = ""
    current_content = []

    for kind, text in elements:
        if kind == "header":
            # Save previous section
            if current_content:
                sections.append((current_header, current_content))
            current_header  = text
            current_content = []
        else:
            current_content.append(text)

    # Don't forget the last section
    if current_content:
        sections.append((current_header, current_content))

    # ── Step 3: Convert sections to chunks ────────────────────
    chunks = []
    MAX_WORDS = 250
    OVERLAP   = 40

    for header, content_lines in sections:
        # Build section text
        if header:
            section_text = header + "\n" + "\n".join(content_lines)
        else:
            section_text = "\n".join(content_lines)

        words = section_text.split()

        if len(words) <= MAX_WORDS:
            # Small section → single chunk
            if len(words) >= 4:
                chunks.append(section_text.strip())
        else:
            # Large section → split with overlap, keeping header on every chunk
            word_list  = words
            start      = 0
            part       = 1
            while start < len(word_list):
                end   = min(start + MAX_WORDS, len(word_list))
                piece = " ".join(word_list[start:end])
                if header and part > 1:
                    piece = f"{header} (continued)\n{piece}"
                if len(piece.split()) >= 4:
                    chunks.append(piece.strip())
                start += MAX_WORDS - OVERLAP
                part  += 1

    return chunks


# ═══════════════════════════════════════════════════════════════
#  DOC  (Legacy .doc — with fallbacks)
# ═══════════════════════════════════════════════════════════════

def parse_doc(filepath):
    # Strategy 1: try as docx (renamed files)
    try:
        return parse_docx(filepath)
    except Exception:
        pass

    # Strategy 2: docx2txt
    try:
        import docx2txt
        text = docx2txt.process(str(filepath))
        if text and text.strip():
            return _chunk_text(text)
    except Exception:
        pass

    # Strategy 3: textract (needs antiword)
    try:
        import textract
        text = textract.process(str(filepath)).decode("utf-8", errors="ignore")
        if text and text.strip():
            return _chunk_text(text)
    except Exception:
        pass

    # Strategy 4: raw ASCII extraction
    try:
        raw     = Path(filepath).read_bytes()
        strings = re.findall(rb"[ -~]{5,}", raw)
        text    = "\n".join(s.decode("ascii", errors="ignore") for s in strings)
        if text.strip():
            return _chunk_text(text)
    except Exception:
        pass

    raise ValueError(
        f"Cannot parse '{Path(filepath).name}'. "
        "Convert to .docx or install antiword: brew install antiword"
    )


# ═══════════════════════════════════════════════════════════════
#  PDF
# ═══════════════════════════════════════════════════════════════

def parse_pdf(filepath):
    from pypdf import PdfReader
    pages = []
    for page in PdfReader(filepath).pages:
        t = page.extract_text()
        if t and t.strip():
            pages.append(t)
    full = "\n\n".join(pages)
    full = re.sub(r"[ \t]+", " ", full)
    full = re.sub(r"\n{3,}", "\n\n", full)
    return _chunk_text(full)


# ═══════════════════════════════════════════════════════════════
#  CSV / TXT
# ═══════════════════════════════════════════════════════════════

def parse_text(filepath):
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        content = f.read()
    return _chunk_text(content)


# ═══════════════════════════════════════════════════════════════
#  Shared word-count chunker (used for PDF/TXT/CSV)
# ═══════════════════════════════════════════════════════════════

def _chunk_text(text, max_words=250, overlap=40):
    """
    Split text at paragraph boundaries into overlapping chunks.
    Used for PDF and plain text where we don't have structural hints.
    """
    chunks, current, cur_len = [], [], 0
    for para in re.split(r"\n\n+", text.strip()):
        para = para.strip()
        if not para:
            continue
        words = para.split()
        if not words:
            continue
        if cur_len + len(words) > max_words and current:
            joined = " ".join(current)
            if len(joined.split()) >= 4:
                chunks.append(joined)
            current = current[-overlap:]
            cur_len = len(current)
        current.extend(words)
        cur_len += len(words)
    if current:
        joined = " ".join(current)
        if len(joined.split()) >= 4:
            chunks.append(joined)
    return chunks


# ═══════════════════════════════════════════════════════════════
#  Public entry point
# ═══════════════════════════════════════════════════════════════

def parse_file(filepath):
    """
    Auto-detect file type and return list of text chunks for the vector DB.
    """
    ext = Path(filepath).suffix.lower()
    if ext in (".xlsx", ".xls"):
        return parse_xlsx(filepath)
    elif ext == ".pdf":
        return parse_pdf(filepath)
    elif ext == ".docx":
        return parse_docx(filepath)
    elif ext == ".doc":
        return parse_doc(filepath)
    elif ext in (".csv", ".txt"):
        return parse_text(filepath)
    else:
        raise ValueError(f"Unsupported file type: '{ext}'. Supported: .xlsx, .pdf, .docx, .doc, .csv, .txt")